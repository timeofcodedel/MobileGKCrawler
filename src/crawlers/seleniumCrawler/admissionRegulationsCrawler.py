import pandas as pd
import xlsxwriter
import time

from filelock import FileLock
from openpyxl.reader.excel import load_workbook
from selenium.common import TimeoutException, WebDriverException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait

# from .baseCrawler import BaseCrawler

from src.crawlers.seleniumCrawler.baseCrawler import BaseCrawler


class SchoolDetail(BaseCrawler):
    def __init__(self):
        super().__init__()

    def schoolSummaryFormGain(self, i, name, schoolsummary):
        lock = FileLock("myfile.lock")
        with lock:
            workbook = load_workbook(r'D:/项目学习/git/MobileGKCrawler/data/招生章程.xlsx')
            sheet = workbook.active
            sheet[f'A{i - 30}'] = name  # type: ignore
            sheet[f'B{i - 30}'] = schoolsummary  # type: ignore
            workbook.save('D:/项目学习/git/MobileGKCrawler/data/招生章程.xlsx')
            workbook.close()
            print(name + 'excel表格录入成功')

    def crawlSchoolOverview(self, numerical_order, name, url):
        self.URL = f'https://www.gaokao.cn/school/{url}/sturule'
        self._chromeDrive.get(self.URL)
        # 找点击的地方
        school_overview = WebDriverWait(self._chromeDrive, 60).until(
            EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR,
                 '#root > div > div.container > div > div > div.layout_layoutBox__2YeRR > div:nth-child(2) > div.main > div > div:nth-child(4) > div > div.clearfix > div.school-content-left_box__2JrKH > div > div:nth-child(2) > div'
                 )
            ))
        school_overview = school_overview[0].text
        # 关闭浏览器窗口
        school_overview = school_overview.split('\n')
        for j in school_overview:
            if '2024年本科招生章程' in j:
                print(school_overview.index(j))
                school_overview = j
                print(school_overview)
                self.schoolSummaryFormGain(numerical_order + 1, name, school_overview)
                self._chromeDrive.close()
                break


class SheetReadout:
    def __init__(self):
        pass

    def gainSchoolUrl(self):
        df = pd.read_excel('D:/项目学习/git/MobileGKCrawler/data/学校.xlsx', sheet_name='Sheet1')
        # print(df.columns)
        URl_data = df['获取地址']
        overallSchoolName = df['学校名字']
        print(URl_data[1], overallSchoolName[1])
        return len(URl_data), URl_data, overallSchoolName

    def createSchoolSummaryForm(self):
        workbook = xlsxwriter.Workbook('D:/项目学习/git/MobileGKCrawler/data/招生章程.xlsx')
        worksheet = workbook.add_worksheet()
        worksheet.write('A1', '学校名字')
        worksheet.write('B1', '招生章程')
        workbook.close()
        print('excel表格创建成功')


if __name__ == '__main__':
    c = SheetReadout().gainSchoolUrl()
    # print(c)
    for i in range(c[0]):
        try:
            bc = SchoolDetail()
            bc.crawlSchoolOverview(i + 31, c[2][i], c[1][i])  # 排序，学校简介，获取地址
        except:
            continue
    pass
    # bc = SchoolDetail()
    # bc.crawlSchoolOverview(31, '北京工业大学', 30)
