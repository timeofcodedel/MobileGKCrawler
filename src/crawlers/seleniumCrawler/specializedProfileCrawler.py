# utf-8
import json
import time

import xlsxwriter
from filelock import FileLock
from openpyxl.reader.excel import load_workbook
from selenium.common import TimeoutException, WebDriverException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from .baseCrawler import BaseCrawler


class SpecializedDetail(BaseCrawler):
    def __init__(self):
        super().__init__()

    def crawlsSpecializedOverview(self, serial_number, url, name):
        self.URL = url
        try:
            self._chromeDrive.get(self.URL)
        except WebDriverException:
            time.sleep(10)
            try:
                self._chromeDrive.get(self.URL)
            except:
                pass

        try:
            expand_button = WebDriverWait(self._chromeDrive, 60).until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR,
                     'em.cursor.colorF60'
                     )
                ))
            expand_button.click()
        except:
            pass

        try:
            specialized_overview = WebDriverWait(self._chromeDrive, 60).until(
                EC.presence_of_all_elements_located(
                    (By.CSS_SELECTOR,
                     '.volspecialty_overviewDes__cAooL'
                     )
                ))
            specialized_overview = specialized_overview[0].text
        except:
            try:
                time.sleep(10)
                specialized_overview = WebDriverWait(self._chromeDrive, 60).until(
                    EC.presence_of_all_elements_located(
                        (By.CSS_SELECTOR,
                         '.volspecialty_overviewDes__cAooL'
                         )
                    ))
                specialized_overview = specialized_overview[0].text
            except:
                specialized_overview = '没有获取到'

        try:
            specialty = WebDriverWait(self._chromeDrive, 60).until(
                EC.presence_of_all_elements_located(
                    (By.CSS_SELECTOR,
                     '.volspecialty_zyxj__Fovtu'
                     )
                ))
            specialty = specialty[0].text
        except:
            try:
                specialty = WebDriverWait(self._chromeDrive, 60).until(
                    EC.presence_of_all_elements_located(
                        (By.CSS_SELECTOR,
                         '.volspecialty_zyxj__Fovtu'
                         )
                    ))
                specialty = specialty[0].text
            except:
                specialty = '没有获取到'

        self.zhuanYeXieRU(serial_number, name, specialized_overview, specialty)
        print(specialized_overview, specialty)
        # 关闭浏览器窗口
        try:
            self._chromeDrive.close()
        except:
            try:
                self._chromeDrive.close()
            except WebDriverException:
                pass

    def zhuanYeXieRU(self, i, name, specializeSummarys, specialtys):  # 专业写入excel表格
        lock = FileLock("myfile.lock")
        with lock:
            workbook = load_workbook(r'./data/专业简介.xlsx')
            sheet = workbook.active
            sheet[f'A{i}'] = name  # type: ignore
            sheet[f'B{i}'] = specializeSummarys  # type: ignore
            sheet[f'C{i}'] = specialtys  # type: ignore
            workbook.save('./data/专业简介.xlsx')
            workbook.close()
            print(name + 'excel表格录入成功')


class SheetReadout:

    def gainSpecializedUrl(self):  # 专业的全部URL
        specializedCrawlerUrls = []
        with open('./data/urls.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
        for category, subcategories in data.items():
            for subcategory, urls in subcategories.items():
                # 将每个子类别中的URL列表添加到all_urls列表中
                specializedCrawlerUrls.extend(urls)
        # print(specializedCrawlerUrls)
        return specializedCrawlerUrls

    def createSpecializedSummaryForm(self):
        workbook = xlsxwriter.Workbook('./data/专业简介.xlsx')
        worksheet = workbook.add_worksheet()
        worksheet.write('A1', '专业名字')
        worksheet.write('B1', '专业简介')
        worksheet.write('C1', '专业详解')
        workbook.close()
        print('excel表格创建成功')


def Launch():
    q = SheetReadout().gainSpecializedUrl()
    SheetReadout().createSpecializedSummaryForm()
    number = 2
    for i in q:
        print(i.split('?special_type=3:'))
        SpecializedDetail().crawlsSpecializedOverview(number, i.split('?special_type=3:')[0],
                                                      i.split('?special_type=3:')[1])
        number += 1
