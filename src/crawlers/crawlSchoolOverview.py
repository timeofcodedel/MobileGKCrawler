from filelock import FileLock
from openpyxl.reader.excel import load_workbook
from selenium.common import TimeoutException, WebDriverException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from src.crawlers.baseCrawler import BaseCrawler
import pandas as pd
import xlsxwriter


class SchoolDetail(BaseCrawler):
    def __init__(self):
        super().__init__()

    def schoolSummaryFormGain(self, i, name, schoolsummary, url):
        lock = FileLock("myfile.lock")
        with lock:
            workbook = load_workbook('../school_overview/学校简介.xlsx')
            sheet = workbook.active
            sheet[f'A{i - 30}'] = name
            sheet[f'B{i - 30}'] = schoolsummary
            sheet[f'C{i - 30}'] = url + '/introDetails'
            workbook.save('学校简介.xlsx')
            workbook.close()
            print(name + 'excel表格录入成功')

    def crawlSchoolOverview(self, numerical_order, name, url):
        try:
            self.URL = f'https://www.gaokao.cn/school/{url}/introDetails'
            self._chromeDrive.get(self.URL)
        except WebDriverException:
            try:
                self._chromeDrive.refresh()
            except WebDriverException:
                try:
                    self._chromeDrive.time.sleep(10)
                    self._chromeDrive.refresh()
                except:
                    pass
        # # 点击详细按钮
        # detail = WebDriverWait(self._chromeDrive, 10).until(
        #     EC.element_to_be_clickable(
        #         (By.CSS_SELECTOR,
        #          ".cursor.school-introduce_more__1vggP"
        #          )
        #     ))
        # detail.click()
        # self._chromeDrive.close()
        # self._chromeDrive.switch_to.window(self._chromeDrive.window_handles[-1])  # 获取全部窗口句柄并且换到最后一个窗口

        # 收集学校介绍
        try:
            school_overview = WebDriverWait(self._chromeDrive, 60).until(
                EC.presence_of_all_elements_located(
                    (By.CSS_SELECTOR,
                     '.richContent'
                     )
                ))
            school_overview = school_overview[0].text
        except TimeoutException:
            try:
                school_overview = WebDriverWait(self._chromeDrive, 60).until(
                    EC.presence_of_all_elements_located(
                        (By.CSS_SELECTOR,
                         '.richContent'
                         )
                    ))
                school_overview = school_overview[0].text
            except:
                school_overview = '无文本内容'

        # 关闭浏览器窗口
        print(numerical_order + 1, name, school_overview, self.URL)
        self.schoolSummaryFormGain(numerical_order + 1, name, school_overview, self.URL)
        self._chromeDrive.close()

    def professionalOpening(self):
        pass


class SheetReadout():
    def __init__(self):
        pass

    def gainSchoolUrl(self):
        df = pd.read_excel('D:/项目学习/git/MobileGKCrawler/data/学校.xlsx', sheet_name='Sheet1')
        # print(df.columns)
        URl_data = df['获取地址']
        overallSchoolName = df['学校名字']
        print(URl_data[1], overallSchoolName[1])
        return len(URl_data), URl_data, overallSchoolName

    def createSchoolSummaryForm(self):
        workbook = xlsxwriter.Workbook('../school_overview/学校简介.xlsx')
        worksheet = workbook.add_worksheet()
        worksheet.write('A1', '学校名字')
        worksheet.write('B1', '学校简介')
        worksheet.write('C1', '获取地址')
        workbook.close()
        print('excel表格创建成功')


if __name__ == "__main__":
    c = SheetReadout().gainSchoolUrl()
    print(c)
    # createSchoolSummaryForm()
    for i in range(c[0]):
        try:
            bc = SchoolDetail()
            bc.crawlSchoolOverview(i + 31, c[2][i], c[1][i])  # 排序，学校简介，获取地址
        except:
            continue
    pass
